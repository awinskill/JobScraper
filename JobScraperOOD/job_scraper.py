from configparser import RawConfigParser
from argparse import ArgumentParser
import sys 
import time
import logging
import openpyxl as xl
from openpyxl.utils import FORMULAE

# Base class for searcher implementations
from abc import ABC, abstractmethod

from concurrent.futures import ThreadPoolExecutor, as_completed


DEV_MODE = True

class Job:
    def __init__(self, id=None, source="LinkedIn", title=None, company=None, location=None, date=None, url=None, salary=None, salary_lower=None, salary_upper=None,description=None, summary=None, fit=None, raw_description=None, cleaned_description=None):
        self.id = id
        self.source = source
        self.url = url
        self.title = title
        self.company = company
        self.location = location
        self.date = date
        self.salary = salary
        self.salary_lower = salary_lower
        self.salary_upper = salary_upper
        self.description = description
        self.summary = summary
        self.fit = fit
        self.raw_description = raw_description
        self.cleaned_description = cleaned_description

    def __str__(self):
        return f"ID: {self.id},Source: {self.source}, Title: {self.title}, Company: {self.company}, Location: {self.location}, \
            Date: {self.date}, URL: {self.url}, Salary: {self.salary}, \
                Salary_Lower: {self.salary_lower}, Salary_Upper: {self.salary_upper}, \
                    Description: {self.description}, Summary: {self.summary}, Fit: {self.fit}"
    def __repr__(self):    
        return f"ID: {self.id},Source: {self.source}, Title: {self.title}, Company: {self.company}, Location: {self.location}, \
            Date: {self.date}, URL: {self.url}, Salary: {self.salary}, \
                Salary_Lower: {self.salary_lower}, Salary_Upper: {self.salary_upper}, \
                    Description: {self.description}, Summary: {self.summary}, Fit: {self.fit}"
    
 # Add this method to convert the object to a dictionary to enable serialization into JSON
    def to_dict(self):
        return {
            "id": self.id,
            "source": self.source,
            "title": self.title,
            "company": self.company,
            "location": self.location,
            "date": self.date,
            "url": self.url,
            "salary": self.salary,
            "salary_lower": self.salary_lower,
            "salary_upper": self.salary_upper,
            "description": self.description,
            "summary": self.summary,
            "fit": self.fit,
            "raw_description": self.raw_description
        }


class ResumeProvider:
    def get_resume(self) -> str:
        return "Your resume text here"




class GenAISummarizer(ABC):
    @abstractmethod
    def summarize(self, job: Job) -> str:
        pass



class SearcherImplementation:
    def __init__(self):
        self.source = ""

    @abstractmethod
    def scrape(self) -> list[Job]:
        # Implement the scraping logic here
        pass

    

# class to sort a job list
class JobSorter(ABC):
    @abstractmethod
    def sort(self, jobs: list[Job]) -> list[Job]:
        #implementation of the sort
        pass

class SalaryDownJobSorter(JobSorter):
    def sort(self, jobs: list[Job]) -> list[Job]:
       sorted_jobs = sorted(jobs, key=lambda job: job.salary_upper or 0, reverse=True)

       return sorted_jobs


# output a job to a file or database
# This is an abstract base class for job writers    
class JobWriter(ABC):
    @abstractmethod
    def write(self, job: Job) -> None:
        # Implement the logic to write the job to a file or database
        pass

class DebugJobWriter(JobWriter):
    def write(self, job: Job) -> None:
        # Implement the logic to write the job to a file or database
        logging.info(f"Writing job to debug output: ")
        logging.info(f"Job title: {job.title}")
        logging.info(f"Company: {job.company}")

class TSVJobWriter(JobWriter):
    def __init__(self, filename: str):
        self.filename = filename
        # write the header to the file
        with open(self.filename, "w") as f:
            f.write("ID\tSource\tTitle\tCompany\tLocation\tDate\tSalary\tSalary_Lower\tSalary_Upper\tSummary\tDescription\tURL\tFit\n")


    def write(self, job: Job) -> None:
        # Implement the logic to write the job to a TSV file
        with open(self.filename, "a") as f:
            f.write(f"{job.id}\t{job.source}\t{job.title}\t{job.company}\t{job.location}\t{job.date}\t{job.salary}\t{job.salary_lower}\t{job.salary_upper}\t{job.summary}\t{job.description}\t{job.url}\t{job.fit}\n")


class XLJobWriter(JobWriter):
    def __init__(self, filename: str):
        logging.info("Initilizing XLJobWriter")

        self.filename = filename
        self.wb=None

        try:
            self.wb = xl.load_workbook(filename=self.filename)

            # The file already exists, so create a new sheet for us to work in
            self.wb.create_sheet()
        except FileNotFoundError:
            # create new workbook
            self.wb = xl.Workbook()

        self.sheet = self.wb.active

        # need to check if the sheet is empty
        if self.sheet.max_row == 1 and self.sheet.max_column == 1:
            # The sheet is empty, so we can write the headings
            self.sheet.title = f"Job Listings - {str(int(time.time()))}"
        else:
            # The sheet is not empty, so we need to create a new sheet
            self.sheet = self.wb.create_sheet(title=f"Job Listings - {int(time.time())}")
        
        # set the column widths
        self.sheet.column_dimensions['A'].width = 15
        self.sheet.column_dimensions['B'].width = 15
        self.sheet.column_dimensions['C'].width = 50
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 15
        self.sheet.column_dimensions['G'].width = 30
        self.sheet.column_dimensions['H'].width = 10
        self.sheet.column_dimensions['I'].width = 10
        self.sheet.column_dimensions['J'].width = 100
        self.sheet.column_dimensions['K'].width = 100
        self.sheet.column_dimensions['L'].width = 50
        self.sheet.column_dimensions['M'].width = 10
        
        # set the column headings

        self.headings = { 
            "ID" : 'A1', 
            "Source":'B1', 
            "Title": 'C1', 
            "Company": 'D1', 
            "Location": 'E1', 
            "Date": 'F1', 
            "Salary": 'G1', 
            "Salary_Lower": 'H1', 
            "Salary_Upper": 'I1', 
            "Summary": 'J1', 
            "Description": 'K1', 
            "URL": 'L1', 
            "Fit": 'M1' 
            }

 
        for key, cell in self.headings.items():
            self.sheet[cell] = key


        # used to keep track of where we are in the sheet
        self.current_row = 2

    # make sure the workbook is saved & closed
    def __del__(self):
        logging.info("Cleaning up XLJobWriter")
        if self.wb is not None:
            self.wb.save(self.filename)
            self.wb.close()


    def write(self, job: Job) -> None:
        
        job_to_column = {
            "id": 1,
            "source": 2,
            "title": 3,
            "company": 4,
            "location": 5,
            "date": 6,
            "salary": 7,
            "salary_lower": 8,
            "salary_upper": 9,
            "summary": 10,
            "description": 11,
            "url": 12,
            "fit": 13
        }

        # Write each attribute of the job to the corresponding column
        for attribute, column in job_to_column.items():
            value = getattr(job, attribute, None)  # Get the value of the attribute
            if isinstance(value, list):
                    value = ", ".join(value)

            self.sheet.cell(row=self.current_row, column=column, value=value)

        # Add a hyperlink formula in column 14 (e.g., "N" column)
        url_cell = f"L{self.current_row}"  # Reference to the URL column
        hyperlink_formula = f'=HYPERLINK({url_cell}, "Link")'
        self.sheet.cell(row=self.current_row, column=14, value=hyperlink_formula)

     #   self.sheet.cell(row=self.current_row, column=14, value=f"=HYPERLINK(M{self.current_row},'link')")

        # Move to the next row
        self.current_row += 1
            



# JobScraper
class JobScraper:

    def __init__(self, scraper: SearcherImplementation):
        self.scraper = scraper
        self.source = scraper.source

        
    def fetch_job_listings(self) -> list[Job]:
        # Use BeautifulSoup or requests to scrape jobs
        # return [Job(title="Software Engineer", url="http://example.com/job1")]
        jobs = self.scraper.scrape()
        return jobs
        
    
    def fetch_job_details(self, job: Job) -> None:
        # Update job.description with full job info
        # Do any local processing of the raw job description here
        logging.debug(f"Fetching job details for {job.title}")



\

# The class that orchestrates the job scraping and summarization process
class BasicJobProcessor():
    def __init__(self, scraper: JobScraper, summarizer: GenAISummarizer, resume_provider: ResumeProvider, job_writer: JobWriter):
        self.scraper = scraper
        self.summarizer = summarizer
        self.resume_provider = resume_provider
        self.job_writer = job_writer

    def process_jobs(self) -> list[Job]:
        resume = self.resume_provider.get_resume()
        jobs = self.scraper.fetch_job_listings()

        job_list = []
        for job in jobs:
            self.scraper.fetch_job_details(job)
            # need to hook up the "resume provider" to the summarizer at some point in the future
            json_job = None
            while json_job is None:
                json_job = self.summarizer.summarize(job.raw_description)
    
            if json_job is None:
                logging.error("Failed to summarize job description")

            job_list.append(json_job)
        
        logging.info(f"Pre-sort job list {len(job_list)}")
           # Filter out None values
        job_list = [job for job in job_list if job is not None]
        logging.info(f"Filtered for null {len(job_list)}")

        job_list = SalaryDownJobSorter().sort(job_list)
        logging.info(f"Post-sort job list {len(job_list)}")

        for job in job_list:
            self.job_writer.write(job)

        return job_list
    

class JobMultiProcessor(BasicJobProcessor):

    def __init__(self, config=RawConfigParser, scrapers=list[JobScraper], summarizer=GenAISummarizer, writers=list[JobWriter], sorter=JobSorter):
        self.scrapers = scrapers
        self.summarizer = summarizer
        self.writers = writers
        self.sorter = sorter
        self.dev_mode = config["DEFAULT"].getboolean("DEV_MODE", fallback=False)
        self.dev_mode_limit = config["DEFAULT"].getint("DEV_MODE_JOB_LIMIT", fallback=2)

    def getScrapers(self):
        return self.scrapers
    def getWriters(self):
        return self.writers
    def getSummarizer(self):
        return self.summarizer
    def getSorter(self):
        return self.sorter
    

    def process_jobs(self) -> list[Job]:
        # 1. Scrape all jobs, iterating over all the scrapers
        # 1.1 Use the supplied AI Summerizer to get summaries for all the jobs
        # 2. Output to all destinations by iterating over all the writers

        job_list = []

        # 1. Scrape all jobs, iterating over all the scrapers
        # Move this to
        for job_scraper in self.scrapers:

            jobs = job_scraper.fetch_job_listings()
            
            if jobs is not None:
                STOP = False
                dev_counter = 0

                for job in jobs:
                    if STOP:
                        break

                    job_scraper.fetch_job_details(job)

                    # Summarize the job
                    json_job = None
                    max_retries = 5
                    retries = 0
                    while json_job is None and retries < max_retries:
                        json_job = self.summarizer.summarize(job)
                        retries += 1

                    if json_job is None:
                        logging.error("Failed to summarize job description after multiple attempts.")
                        continue

                    job_list.append(json_job)
                    dev_counter += 1

                    if self.dev_mode and dev_counter > self.dev_mode_limit:
                        logging.info(f"Dev mode: stopping after {self.dev_mode_limit} jobs")
                        STOP = True
                        break
            else:
                logging.warning("No jobs found.")
                return job_list
                        
        logging.info(f"Pre-sort job list {len(job_list)}")
        # Filter out None values
        job_list = [job for job in job_list if job is not None]
        logging.info(f"Filtered for null {len(job_list)}")

        job_list = self.sorter.sort(job_list)
        logging.info(f"Post-sort job list {len(job_list)}")

# 3. Write to all destinations
        for job_writer in self.writers:
            for job in job_list:
                job_writer.write(job)

        return job_list


class JobMultiParallelProcessor(JobMultiProcessor):

    def __init__(self, config=RawConfigParser, scrapers=list[JobScraper], summarizer=GenAISummarizer, writers=list[JobWriter], sorter=JobSorter):
        super().__init__(config, scrapers, summarizer, writers, sorter)
        self.executor = ThreadPoolExecutor(len(scrapers))
        self.futures = []
        self.job_list = []
        self.max_workers = config["DEFAULT"].getint("MAX_WORKERS", fallback=5)


    def process_jobs(self) -> list[Job]:
        job_list = []

        # Use ThreadPoolExecutor to run scrapers in parallel
        with ThreadPoolExecutor(max_workers=len(self.scrapers)) as executor:
            future_to_scraper = {executor.submit(scraper.fetch_job_listings): scraper for scraper in self.scrapers}

            for future in as_completed(future_to_scraper):
                scraper = future_to_scraper[future]
                try:
                    jobs = future.result()  # Get the result of the scraper
                    if jobs is not None:
                        logging.info(f"{scraper.source}: Found {len(jobs)} jobs.")
                        job_list.extend(self._process_jobs_from_scraper(scraper, jobs))
                    else:
                        logging.warning(f"{scraper.source}: No jobs found.")
                except Exception as e:
                    logging.error(f"{scraper.source}: Failed to fetch job listings. Error: {e}")

        logging.info(f"Pre-sort job list {len(job_list)}")
        # Filter out None values
        job_list = [job for job in job_list if job is not None]
        logging.info(f"Filtered for null {len(job_list)}")

        # Sort the jobs
        job_list = self.sorter.sort(job_list)
        logging.info(f"Post-sort job list {len(job_list)}")

        # Write to all destinations
        for job_writer in self.writers:
            for job in job_list:
                job_writer.write(job)

        return job_list

    def _process_jobs_from_scraper(self, scraper, jobs):
        """Process jobs from a single scraper."""
        processed_jobs = []
        STOP = False
        dev_counter = 0

        for job in jobs:
            if STOP:
                break

            scraper.fetch_job_details(job)

            # Summarize the job
            json_job = None
            max_retries = 5
            retries = 0
            while json_job is None and retries < max_retries:
                json_job = self.summarizer.summarize(job)
                retries += 1

            if json_job is None:
                logging.error("Failed to summarize job description after multiple attempts.")
                continue

            processed_jobs.append(json_job)
            dev_counter += 1

            if self.dev_mode and dev_counter >= self.dev_mode_limit:
                logging.info(f"Dev mode: stopping after {self.dev_mode_limit} jobs")
                STOP = True
                break

        return processed_jobs



if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
    logging.info("Starting Job Scraper test framework...")
  